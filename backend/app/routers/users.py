"""Users router — admin user management with single + bulk creation."""

import csv
import io
import logging
import uuid

from fastapi import APIRouter, Depends, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.dependencies import get_current_admin
from app.models.database import User, get_db
from app.models.schemas import (
    AdminResetPasswordRequest,
    BulkCreateResponse,
    UserCreateRequest,
    UserListResponse,
    UserResponse,
    UserUpdateRequest,
)
from app.services.audit_service import AuditService
from app.services.auth_service import hash_password

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1/users", tags=["users"])
audit_service = AuditService()


@router.get("", response_model=UserListResponse)
async def list_users(
    search: str | None = None,
    role: str | None = None,
    is_active: bool | None = None,
    limit: int = 50,
    offset: int = 0,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_admin),
):
    """List users with pagination, search, and filters."""
    limit = min(limit, 200)
    filters = []
    if search:
        filters.append(
            (User.name.ilike(f"%{search}%")) | (User.email.ilike(f"%{search}%"))
        )
    if role:
        filters.append(User.role == role)
    if is_active is not None:
        filters.append(User.is_active == is_active)

    count_stmt = select(func.count(User.id)).where(*filters)
    total = (await db.execute(count_stmt)).scalar() or 0

    stmt = (
        select(User)
        .where(*filters)
        .order_by(User.created_at.desc())
        .limit(limit)
        .offset(offset)
    )
    result = await db.execute(stmt)
    users = list(result.scalars().all())

    return UserListResponse(
        users=[UserResponse.model_validate(u) for u in users],
        total=total,
        limit=limit,
        offset=offset,
    )


@router.post("", response_model=UserResponse, status_code=201)
async def create_user(
    body: UserCreateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_admin),
):
    """Create a single user."""
    if body.role not in ("admin", "manager", "user"):
        raise HTTPException(status_code=400, detail="Rôle invalide")

    existing = await db.execute(select(User).where(User.email == body.email))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Cet email est déjà utilisé")

    existing_mat = await db.execute(select(User).where(User.matricule == body.matricule))
    if existing_mat.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Ce matricule est déjà utilisé")

    user = User(
        email=body.email,
        name=body.name,
        matricule=body.matricule,
        role=body.role,
        password_hash=hash_password(body.password),
        is_active=True,
    )
    db.add(user)
    await db.flush()

    await audit_service.log(
        db=db,
        user_id=current_user.id,
        action="user_created",
        entity_type="user",
        entity_id=user.id,
        details={"email": body.email, "role": body.role},
    )

    await db.commit()
    await db.refresh(user)
    return UserResponse.model_validate(user)


@router.put("/{user_id}", response_model=UserResponse)
async def update_user(
    user_id: uuid.UUID,
    body: UserUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_admin),
):
    """Update user details."""
    user = await db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur introuvable")

    details = {}
    if body.name is not None:
        details["old_name"] = user.name
        user.name = body.name
        details["new_name"] = body.name
    if body.email is not None:
        existing = await db.execute(
            select(User).where(User.email == body.email, User.id != user_id)
        )
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=409, detail="Cet email est déjà utilisé")
        details["old_email"] = user.email
        user.email = body.email
        details["new_email"] = body.email
    if body.matricule is not None:
        existing_mat = await db.execute(
            select(User).where(User.matricule == body.matricule, User.id != user_id)
        )
        if existing_mat.scalar_one_or_none():
            raise HTTPException(status_code=409, detail="Ce matricule est déjà utilisé")
        details["old_matricule"] = user.matricule
        user.matricule = body.matricule
        details["new_matricule"] = body.matricule
    if body.role is not None:
        if body.role not in ("admin", "manager", "user"):
            raise HTTPException(status_code=400, detail="Rôle invalide")
        details["old_role"] = user.role
        user.role = body.role
        details["new_role"] = body.role
    if body.is_active is not None:
        details["old_is_active"] = user.is_active
        user.is_active = body.is_active
        details["new_is_active"] = body.is_active

    await audit_service.log(
        db=db,
        user_id=current_user.id,
        action="user_updated",
        entity_type="user",
        entity_id=user_id,
        details=details,
    )

    await db.commit()
    await db.refresh(user)
    return UserResponse.model_validate(user)


@router.delete("/{user_id}", status_code=204)
async def delete_user(
    user_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_admin),
):
    """Deactivate a user (prevent self-deletion)."""
    if user_id == current_user.id:
        raise HTTPException(
            status_code=400,
            detail="Impossible de supprimer votre propre compte",
        )

    user = await db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur introuvable")

    user.is_active = False

    await audit_service.log(
        db=db,
        user_id=current_user.id,
        action="user_deactivated",
        entity_type="user",
        entity_id=user_id,
        details={"email": user.email},
    )

    await db.commit()


@router.post("/{user_id}/reset-password", status_code=200)
async def reset_user_password(
    user_id: uuid.UUID,
    body: AdminResetPasswordRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_admin),
):
    """Admin resets a user's password."""
    user = await db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur introuvable")

    user.password_hash = hash_password(body.new_password)

    await audit_service.log(
        db=db,
        user_id=current_user.id,
        action="user_password_reset",
        entity_type="user",
        entity_id=user_id,
        details={"email": user.email},
    )

    await db.commit()
    return {"message": "Mot de passe réinitialisé avec succès"}


@router.post("/bulk", response_model=BulkCreateResponse)
async def bulk_create_users(
    file: UploadFile,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_admin),
):
    """Bulk create users from CSV or XLSX file."""
    filename = file.filename or ""
    content = await file.read()

    rows: list[dict] = []

    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append(row)
    elif filename.endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="Fichier XLSX vide")
        headers = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = str(val or "").strip()
            if any(row_dict.values()):
                rows.append(row_dict)
    else:
        raise HTTPException(
            status_code=400,
            detail="Format non supporté. Utilisez CSV ou XLSX.",
        )

    created: list[UserResponse] = []
    errors: list[dict] = []

    for i, row in enumerate(rows):
        email = row.get("email", "").strip()
        name = row.get("nom", "").strip() or row.get("name", "").strip()
        matricule = row.get("matricule", "").strip()
        password = row.get("mot_de_passe", "").strip() or row.get("password", "").strip()
        role = row.get("role", "user").strip().lower()

        line_num = i + 2  # 1-indexed + header

        if not email:
            errors.append({"line": line_num, "error": "Email manquant"})
            continue
        if not name:
            errors.append({"line": line_num, "error": "Nom manquant", "email": email})
            continue
        if not matricule:
            errors.append({"line": line_num, "error": "Matricule manquant", "email": email})
            continue
        if not password:
            errors.append({"line": line_num, "error": "Mot de passe manquant", "email": email})
            continue
        if role not in ("admin", "manager", "user"):
            role = "user"

        # Check duplicate email
        existing = await db.execute(select(User).where(User.email == email))
        if existing.scalar_one_or_none():
            errors.append({"line": line_num, "error": "Email déjà utilisé", "email": email})
            continue

        # Check duplicate matricule
        existing_mat = await db.execute(select(User).where(User.matricule == matricule))
        if existing_mat.scalar_one_or_none():
            errors.append({"line": line_num, "error": "Matricule déjà utilisé", "email": email})
            continue

        user = User(
            email=email,
            name=name,
            matricule=matricule,
            role=role,
            password_hash=hash_password(password),
            is_active=True,
        )
        db.add(user)
        await db.flush()
        created.append(UserResponse.model_validate(user))

    if created:
        await audit_service.log(
            db=db,
            user_id=current_user.id,
            action="users_bulk_created",
            entity_type="user",
            details={"count": len(created), "errors": len(errors)},
        )
        await db.commit()

    return BulkCreateResponse(created=created, errors=errors)


@router.get("/bulk-template")
async def download_bulk_template(
    current_user: User = Depends(get_current_admin),
):
    """Download CSV template for bulk user creation."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["email", "nom", "matricule", "mot_de_passe", "role"])
    writer.writerow(["exemple@beac.int", "Jean Dupont", "MAT001", "MotDePasse123!", "user"])
    content = output.getvalue()

    return StreamingResponse(
        io.BytesIO(content.encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=modele_utilisateurs.csv"},
    )


@router.get("/bulk-template-xlsx")
async def download_bulk_template_xlsx(
    current_user: User = Depends(get_current_admin),
):
    """Download XLSX template for bulk user creation."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Utilisateurs"
    ws.append(["email", "nom", "matricule", "mot_de_passe", "role"])
    ws.append(["exemple@beac.int", "Jean Dupont", "MAT001", "MotDePasse123!", "user"])

    # Auto-size columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modele_utilisateurs.xlsx"},
    )
